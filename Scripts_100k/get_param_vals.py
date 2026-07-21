from __future__ import annotations
# import required python modules
import pandas as pd
import openpyxl
from dataclasses import dataclass
from typing import Dict, Iterator, Tuple, Any, Optional

class ParamValues:
    def __init__(self, excel_file):
        self.excel_file = excel_file

    def get_param_list(self):
        fc_dict = {}
        # Read Excel Data
        excel_data = pd.read_excel(self.excel_file, sheet_name=None, keep_default_na=False)
        
        # Data Prep Clean
        fc_to_create_buffer_zone = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Function Name"] == "FC to create buffer zone", "FeatureClass"])
        fc_dict['buffer_points_25K'] = fc_to_create_buffer_zone
        not_include_fields = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "Not include fields", "Field"])
        fc_dict['not_include_fields'] = not_include_fields
        feature_to_split = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "Feature_to_split", "FeatureClass"])
        fc_dict['feature_to_split'] = feature_to_split
        bau_field_fc = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "BAU Field FC", "FeatureClass"])
        fc_dict['bau_field_fc'] = bau_field_fc

        # Transportation
        trans_build_up_buildings = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['trans_build_up_buildings'] = trans_build_up_buildings
        topology_features = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Feature Usage Notes"] == "Topology Features", "FeatureClass"])
        fc_dict['topology_features'] = topology_features
        collapse_sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Collapse Road Detail & Replace", "Query"])
        fc_dict['collapse_sql'] = collapse_sql
        railway_sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Convert close single track to double", "Expression"])
        fc_dict['railway_sql'] = railway_sql
        fcs_trim_extent_trans = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "DATA_TRIM_EXTEND", "FeatureClass"])
        fc_dict['fcs_trim_extent_trans'] = fcs_trim_extent_trans
        sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Feature to Point", "Query"])
        fc_dict['sql'] = sql

        # Hydrography
        fcs_trim_extent_hyd = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "DATA_TRIM_EXTEND", "FeatureClass"])
        fc_dict['fcs_trim_extent_hyd'] = fcs_trim_extent_hyd
        hydro_prep_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Hydro Prep", "FeatureClass"])
        fc_dict['hydro_prep_fc_list'] = hydro_prep_fc_list
        hydro_input_polygon_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Input polygon Feature Class", "FeatureClass"])
        fc_dict['hydro_input_polygon_fc'] = hydro_input_polygon_fc
        hydro_center_line_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Centerlines Feature Class", "FeatureClass"])
        fc_dict['hydro_center_line_fc'] = hydro_center_line_fc
        hydro_replace_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Replace Feature Class", "FeatureClass"])
        fc_dict['hydro_replace_fc'] = hydro_replace_fc
        hydro_remove_near_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Remove Near Polygons", "FeatureClass"])
        fc_dict['hydro_remove_near_poly_list'] = hydro_remove_near_poly_list
        hydro_enlarge_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Enlarge Polygons", "FeatureClass"])
        fc_dict['hydro_enlarge_poly_list'] = hydro_enlarge_poly_list
        hydro_remove_small_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Remove Small Polygons by Converting", "FeatureClass"])
        fc_dict['hydro_remove_small_poly_list'] = hydro_remove_small_poly_list
        hydro_erase_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Erase Polygons", "FeatureClass"])
        fc_dict['hydro_erase_poly_list'] = hydro_erase_poly_list
        hydro_small_line_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Feature 2 Point Line", "FeatureClass"])
        fc_dict['hydro_small_line_fc_list'] = hydro_small_line_fc_list
        hydro_small_point_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Feature 2 Point Point", "FeatureClass"])
        fc_dict['hydro_small_point_fc_list'] = hydro_small_point_fc_list
        hydro_delete_small_pools = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Delete Small Pools", "FeatureClass"])
        fc_dict['hydro_delete_small_pools'] = hydro_delete_small_pools

        # Built-up Generalization
        cut_build_a_inputs_poly = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Cut Polygons by Road and Track", "FeatureClass"])
        fc_dict['cut_build_a_inputs_poly'] = cut_build_a_inputs_poly
        small_bldg_2_point_a = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Polygon", "FeatureClass"])
        fc_dict['small_bldg_2_point_a'] = small_bldg_2_point_a
        small_bldg_2_point_p = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Point", "FeatureClass"])
        fc_dict['small_bldg_2_point_p'] = small_bldg_2_point_p
        features_in_cemetery = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Buildings in Cemetery", "FeatureClass"])
        fc_dict['features_in_cemetery'] = features_in_cemetery
        enlarge_barrier_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Enlarge Builtup Features (Cemetery)", "FeatureClass"])
        fc_dict['enlarge_barrier_features'] = enlarge_barrier_features
        delete_small_bldgs = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Small Buildings", "FeatureClass"])
        fc_dict['delete_small_bldgs'] = delete_small_bldgs
        enlarge_building_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Enlarge Small Buildings", "FeatureClass"])
        fc_dict['enlarge_building_features'] = enlarge_building_features
        delineate_building_layers = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Input Building Layers", "FeatureClass"])
        fc_dict['delineate_building_layers'] = delineate_building_layers
        delineate_edge_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Edge Features", "FeatureClass"])
        fc_dict['delineate_edge_features'] = delineate_edge_features
        delete_small_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Small Features (Swimming)", "FeatureClass"])
        fc_dict['delete_small_features'] = delete_small_features

        # Vegetation
        veg_lyrs_list = list(excel_data["7_Vegetation"].loc[excel_data["7_Vegetation"]["Function Name"] == "Vegetation", "FeatureClass"])
        fc_dict['veg_lyrs_list'] = veg_lyrs_list
        veg_field_values = list(excel_data["7_Vegetation"].loc[excel_data["7_Vegetation"]["Rule Type"] == "Calculate", "Value"])
        fc_dict['veg_field_values'] = veg_field_values
        veg_transfer_veg_features = dict(zip(excel_data["7_Vegetation"]
                .loc[excel_data["7_Vegetation"]["Function Name"].eq("Transfer Feature"),"FeatureClass"],
                excel_data["7_Vegetation"].loc[excel_data["7_Vegetation"]["Function Name"].eq("Transfer Feature"),"Feature Usage Notes"]
            )
        )
        fc_dict['veg_transfer_veg_features'] = veg_transfer_veg_features

        # Utility
        utility_area_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Polygon", "FeatureClass"])
        fc_dict['utility_area_features'] = utility_area_features
        utility_point_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Point", "FeatureClass"])
        fc_dict['utility_point_features'] = utility_point_features
        utility_compare_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['utility_compare_features'] = utility_compare_features
        utility_merge_clusters = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Merge Clusters", "FeatureClass"])
        fc_dict['utility_merge_clusters'] = utility_merge_clusters

        # Hypsography
        hypso_compare_features = list(excel_data["6_Hypsography"].loc[excel_data["6_Hypsography"]["Function Name"] == "Erase Vegetation Hypso", "FeatureClass"])
        fc_dict['hypso_compare_features'] = hypso_compare_features

        # Detect conflict
        Structure2Structure = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Structure2Structure", "FeatureClass"])
        fc_dict['Structure2Structure'] = Structure2Structure
        Structure2Lines = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Structure2Lines", "FeatureClass"])
        fc_dict['Structure2Lines'] = Structure2Lines
        Lines2Lines = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Lines2Lines", "FeatureClass"])
        fc_dict['Lines2Lines'] = Lines2Lines
        G1_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G1_Poly2Poly", "FeatureClass"])
        fc_dict['G1_Poly2Poly'] = G1_Poly2Poly
        G2_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G2_Poly2Poly", "FeatureClass"])
        fc_dict['G2_Poly2Poly'] = G2_Poly2Poly
        G3_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G3_Poly2Poly", "FeatureClass"])
        fc_dict['G3_Poly2Poly'] = G3_Poly2Poly
        df_query_input_lyr = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Apply_layer_definition", "FeatureClass"])
        fc_dict['df_query_input_lyr'] = df_query_input_lyr

        # Resolve conflicts buildings
        built_up_area_fcs = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Function Name"] == "Hide Buildings Under Generalized and Town Built-Up Area", "FeatureClass"])
        fc_dict['build_up_area_fcs'] = built_up_area_fcs
        input_building_layers = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Input Building Layers", "FeatureClass"])
        fc_dict['input_building_layers'] = input_building_layers
        input_barrier_layers = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Input Barrier Layers", "FeatureClass"])
        fc_dict['input_barrier_layers'] = input_barrier_layers
        g5_input_points = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G5 Input Points", "FeatureClass"])
        fc_dict['g5_input_points'] = g5_input_points
        g7_input_points = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G7 Input Points", "FeatureClass"])
        fc_dict['g7_input_points'] = g7_input_points
        g1_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G1 Align Features", "FeatureClass"])
        fc_dict['g1_align_features'] = g1_align_features
        g4_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G4 Align Features", "FeatureClass"])
        fc_dict['g4_align_features'] = g4_align_features
        g5_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G5 Align Features", "FeatureClass"])
        fc_dict['g5_align_features'] = g5_align_features
        g6_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G6 Align Features", "FeatureClass"])
        fc_dict['g6_align_features'] = g6_align_features
        g7_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G7 Align Features", "FeatureClass"])
        fc_dict['g7_align_features'] = g7_align_features
        input_primary = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Convert Polygons", "FeatureClass"])
        fc_dict['input_primary'] = input_primary
        input_secondary = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Compare features", "FeatureClass"])
        fc_dict['input_secondary'] = input_secondary
        df_query_input_lyr_9b = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Function Name"] == "Apply_layer_definition", "FeatureClass"])
        fc_dict['df_query_input_lyr_9b'] = df_query_input_lyr_9b

        # Resolve conflicts lines
        input_line_layers = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Input Line Layers", "FeatureClass"])
        fc_dict['input_line_layers'] = input_line_layers
        edge_features = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Edge Feature", "FeatureClass"])
        fc_dict['edge_features'] = edge_features
        embank_list = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Embankment Features", "FeatureClass"])
        fc_dict['embank_list'] = embank_list
        compare_fcs_embank = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['compare_fcs_embank'] = compare_fcs_embank
        road_query = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "road_query", "Expression"])
        fc_dict['road_query'] = road_query
        bridge_query = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "bridge_query", "Expression"])
        fc_dict['bridge_query'] = bridge_query
        footprint_fcs = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Function Name"] == "Snap Points to Shifted Lines", "FeatureClass"])
        fc_dict['footprint_fcs'] = footprint_fcs
        resolve_line_compare = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Function Name"] == "Resolve Conflicts for Lakes and Ponds", "FeatureClass"])
        fc_dict['resolve_line_compare'] = resolve_line_compare

        # Apply carto symbology
        attribution_fc_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "FeatureClass"])
        fc_dict['attribution_fc_list'] = attribution_fc_list
        express_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Expression"])
        fc_dict['express_list'] = express_list
        query_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Query"])
        fc_dict['query_list'] = query_list
        field_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Field"])
        fc_dict['field_list'] = field_list
        intersecting_fc_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Split Embankments and Cuttings", "FeatureClass"])
        fc_dict['intersecting_fc_list'] = intersecting_fc_list
        prep_line_resolve_fcs_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Prep For Line Resolve", "FeatureClass"])
        fc_dict['prep_line_resolve_fcs_list'] = prep_line_resolve_fcs_list
        apply_symbology_layers_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Calculate VST on Workspace", "FeatureClass"])
        fc_dict['apply_symbology_layers_list'] = apply_symbology_layers_list

        align_bridge_point_all = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Align Bridge Point", "FeatureClass"])
        fc_dict['align_bridge_point_input'] = str(align_bridge_point_all[0]) if align_bridge_point_all else None
        
        align_bridge_point_waterbody = list(excel_data["8_ApplyCartoSymbology"].loc[
            (excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Align Bridge Point") & 
            (excel_data["8_ApplyCartoSymbology"]["Feature Usage Notes"].str.contains("Waterbody", na=False, case=False)), "FeatureClass"])
        fc_dict['align_bridge_point_waterbody'] = align_bridge_point_waterbody
        
        align_bridge_point_surface = list(excel_data["8_ApplyCartoSymbology"].loc[
            (excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Align Bridge Point") & 
            (excel_data["8_ApplyCartoSymbology"]["Feature Usage Notes"].str.contains("SurfaceLine", na=False, case=False)), "FeatureClass"])
        fc_dict['align_bridge_point_surface'] = align_bridge_point_surface
        return fc_dict
    
    def get_param_vals(self):
        sheet_list = ['1_DataPreparation', '2_Transportation', '3_Hydrography', '4_Built Environment',
                    '5_Utility', '6_Hypsography', '7_Vegetation', '8_ApplyCartoSymbology', '9a_ResolveConflictsLines',
                    '9b_ResolveConflictsBuildings', '10_DetectConflicts']
        # Load workbook and select sheet
        val_dict = {}
        wb = openpyxl.load_workbook(self.excel_file, data_only=True)
        for sheet_name in sheet_list:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            ws = wb[sheet_name]
            # Iterate row by row
            for row in ws.iter_rows(values_only=True):
                for i, cell_value in enumerate(row):
                    if cell_value is not None and str(cell_value).strip():
                        # If the key value in empty in excell then assign ""
                        next_val = row[i + 1] if i + 1 < len(row) else ""
                        if next_val is None:
                            next_val = ""
                        val_dict[str(cell_value).strip()] = next_val

        return val_dict


@dataclass
class LayerNames:
    _map: Dict[str, Any]

    def __getattr__(self, name: str) -> Any:
        try:
            return self._map[name]
        except KeyError as e:
            raise AttributeError(f"No such layer key: {name!r}") from e

    def __getitem__(self, key: str) -> Any:
        return self._map[key]

    def get(self, key: str, default: Any = None) -> Any:
        return self._map.get(key, default)

    def keys(self):
        return self._map.keys()

    def items(self):
        return self._map.items()

    def __iter__(self) -> Iterator[str]:
        return iter(self._map)

    def __len__(self) -> int:
        return len(self._map)


def _is_nan(x: Any) -> bool:
    return pd.isna(x)


def _as_key(x: Any) -> Optional[str]:
    if x is None or _is_nan(x):
        return None
    s = str(x).strip()
    return s or None


class Validator:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file

    def get_layer_names(self, excel_sheet_name: str) -> LayerNames:
        df = pd.read_excel(
            self.excel_file,
            sheet_name=excel_sheet_name,
            usecols="A:B",          # only columns A, B
            engine="openpyxl",
            header=0,
            dtype=object,
        )

        mapping: Dict[str, Any] = {}

        for a, b in df.itertuples(index=False, name=None):
            key_a = _as_key(a)
            val_b = None if _is_nan(b) else b

            # skip empty rows
            if key_a is None and val_b is None:
                continue

            if not key_a:
                continue

            # Optional: protect against conflicting duplicates
            if key_a in mapping and mapping[key_a] != val_b:
                raise ValueError(f"Duplicate/conflicting key {key_a!r}: {mapping[key_a]!r} vs {val_b!r}")

            mapping[key_a] = val_b

        return LayerNames(mapping)