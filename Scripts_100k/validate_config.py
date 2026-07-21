"""Pre-run validation of the GenCarto100k Excel configuration file.

Checks, without needing arcpy or ArcGIS Pro:
  1. All sheets expected by get_param_vals are present.
  2. Every val_dict key that main.py reads exists in the workbook, so a
     renamed or shifted label fails here instead of mid-run.
  3. Keys defined in more than one cell with different values (the flattening
     in get_param_vals silently keeps the last one).

Usage:
    python validate_config.py <path-to-config.xlsx>

Exits with code 1 if any problem is found.
"""
import re
import sys
import os

import openpyxl

SHEETS = ['1_DataPreparation', '2_Transportation', '3_Hydrography', '4_Built Environment',
          '5_Utility', '6_Hypsography', '7_Vegetation', '8_ApplyCartoSymbology',
          '9a_ResolveConflictsLines', '9b_ResolveConflictsBuildings', '10_DetectConflicts',
          '11_LoadDataFinal100K']


def flatten_workbook(xlsx_path):
    """Replicates ParamValues.get_param_vals: every non-empty cell is a key
    whose value is the cell to its right. Also records where each key came from."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    val_dict = {}
    key_origin = {}
    missing_sheets = [s for s in SHEETS if s not in wb.sheetnames]
    for sheet_name in SHEETS:
        if sheet_name in missing_sheets:
            continue
        ws = wb[sheet_name]
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for i, cell_value in enumerate(row):
                if cell_value is not None and str(cell_value).strip():
                    key = str(cell_value).strip()
                    next_val = row[i + 1] if i + 1 < len(row) else ""
                    if next_val is None:
                        next_val = ""
                    key_origin.setdefault(key, []).append((sheet_name, r, i + 1, next_val))
                    val_dict[key] = next_val
    return val_dict, key_origin, missing_sheets


def keys_read_by_main(main_path):
    """All val_dict['...'] keys referenced by non-comment lines of main.py."""
    keys = []
    with open(main_path, encoding="utf-8") as f:
        for line in f:
            if line.lstrip().startswith("#"):
                continue
            keys.extend(re.findall(r"val_dict\[['\"]([^'\"]+)['\"]\]", line))
    return keys


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        return 1
    xlsx_path = sys.argv[1]
    if not os.path.isfile(xlsx_path):
        print(f"ERROR: config file not found: {xlsx_path}")
        return 1
    main_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "main.py")

    problems = 0
    val_dict, key_origin, missing_sheets = flatten_workbook(xlsx_path)

    for sheet in missing_sheets:
        print(f"ERROR: sheet missing from workbook: {sheet}")
        problems += 1

    used = keys_read_by_main(main_path)
    print(f"main.py reads {len(used)} configuration keys")
    for key in used:
        if key not in val_dict:
            print(f"ERROR: key read by main.py but not found in workbook: {key}")
            problems += 1

    for key in used:
        origins = key_origin.get(key, [])
        if len(origins) > 1 and len({repr(v) for (_, _, _, v) in origins}) > 1:
            locations = "; ".join(f"{s} row {r} col {c} -> {v!r}" for (s, r, c, v) in origins)
            print(f"WARNING: key defined in multiple cells with different values "
                  f"(the last one wins): {key}: {locations}")
            problems += 1

    if problems:
        print(f"\n{problems} problem(s) found.")
        return 1
    print("Configuration OK.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
